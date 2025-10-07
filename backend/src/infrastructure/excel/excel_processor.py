"""
Excel File Processor

Handles Excel file parsing, multi-sheet support, and initial data extraction.
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Any
from pathlib import Path
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Processes Excel files and extracts metadata."""
    
    SUPPORTED_FORMATS = ['.xlsx', '.xls', '.xlsm']
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
    
    def __init__(self):
        self.current_file = None
        self.workbook = None
    
    async def process_file(self, file_path: str, user_id: int) -> Dict[str, Any]:
        """
        Process an Excel file and extract all metadata.
        
        Args:
            file_path: Path to the Excel file
            user_id: ID of the user uploading the file
            
        Returns:
            Dictionary containing file metadata and sheet information
        """
        try:
            path = Path(file_path)
            
            # Validate file
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if path.suffix.lower() not in self.SUPPORTED_FORMATS:
                raise ValueError(f"Unsupported file format: {path.suffix}")
            
            file_size = path.stat().st_size
            if file_size > self.MAX_FILE_SIZE:
                raise ValueError(f"File too large: {file_size / 1024 / 1024:.2f} MB (max {self.MAX_FILE_SIZE / 1024 / 1024} MB)")
            
            logger.info(f"Processing Excel file: {path.name} ({file_size / 1024:.2f} KB)")
            
            # Load workbook
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract metadata
            metadata = {
                'filename': path.name,
                'file_size': file_size,
                'user_id': user_id,
                'sheet_names': self.workbook.sheetnames,
                'sheet_count': len(self.workbook.sheetnames),
                'created_at': datetime.utcnow().isoformat(),
                'sheets': []
            }
            
            # Process each sheet
            total_rows = 0
            total_columns = 0
            for index, sheet_name in enumerate(self.workbook.sheetnames):
                sheet_data = await self.process_sheet(file_path, sheet_name, index)
                metadata['sheets'].append(sheet_data)
                total_rows += sheet_data.get('row_count', 0)
                total_columns = max(total_columns, sheet_data.get('column_count', 0))
            
            # Add totals to metadata
            metadata['total_rows'] = total_rows
            metadata['total_columns'] = total_columns
            
            logger.info(f"Successfully processed {metadata['sheet_count']} sheets")
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise
    
    async def process_sheet(self, file_path: str, sheet_name: str, sheet_index: int = 0) -> Dict[str, Any]:
        """
        Process a single sheet from the Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to process
            sheet_index: Index of the sheet in the workbook
            
        Returns:
            Dictionary containing sheet metadata
        """
        try:
            # Read sheet with pandas
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Basic metadata
            sheet_data = {
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'row_count': len(df),
                'column_count': len(df.columns),
                'columns': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'has_header': self._detect_header(df),
                'memory_usage': int(df.memory_usage(deep=True).sum())
            }
            
            logger.info(f"Processed sheet '{sheet_name}': {sheet_data['row_count']} rows, {sheet_data['column_count']} columns")
            
            return sheet_data
            
        except Exception as e:
            logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
            raise
    
    async def get_sheet_data(
        self, 
        file_path: str, 
        sheet_name: str,
        start_row: int = 0,
        max_rows: int = 100
    ) -> pd.DataFrame:
        """
        Get data from a specific sheet with pagination.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row index
            max_rows: Maximum number of rows to return
            
        Returns:
            DataFrame containing the requested data
        """
        try:
            # Read with skiprows and nrows for pagination
            df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name,
                skiprows=start_row if start_row > 0 else None,
                nrows=max_rows
            )
            
            return df
            
        except Exception as e:
            logger.error(f"Error reading sheet data: {str(e)}")
            raise
    
    async def get_sheet_preview(
        self, 
        file_path: str, 
        sheet_name: str,
        rows: int = 100
    ) -> Dict[str, Any]:
        """
        Get a preview of sheet data with formatted output.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            rows: Number of rows to preview
            
        Returns:
            Dictionary with preview data
        """
        try:
            df = await self.get_sheet_data(file_path, sheet_name, max_rows=rows)
            
            # Convert to serializable format
            preview = {
                'sheet_name': sheet_name,
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'data': df.fillna('').to_dict(orient='records'),
                'preview_size': min(rows, len(df))
            }
            
            return preview
            
        except Exception as e:
            logger.error(f"Error getting sheet preview: {str(e)}")
            raise
    
    def _detect_header(self, df: pd.DataFrame) -> bool:
        """
        Detect if the DataFrame has a proper header row.
        
        Args:
            df: DataFrame to check
            
        Returns:
            True if header is detected, False otherwise
        """
        # Heuristic: Check if first row has different types than rest
        if len(df) == 0:
            return False
        
        # If all columns are named like 'Unnamed: X', no header
        unnamed_count = sum(1 for col in df.columns if str(col).startswith('Unnamed:'))
        if unnamed_count > len(df.columns) / 2:
            return False
        
        return True
    
    def close(self):
        """Close the workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
